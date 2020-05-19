# Copyright (c) 2020 Cisco and/or its affiliates.
#
# This software is licensed to you under the terms of the Cisco Sample
# Code License, Version 1.1 (the "License"). You may obtain a copy of the
# License at
#
#                https://developer.cisco.com/docs/licenses
#
# All use of the material herein must be in accordance with the terms of
# the License. All rights not expressly granted by the License are
# reserved. Unless required by applicable law or agreed to separately in
# writing, software distributed under the License is distributed on an "AS
# IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
# or implied.
 
import argparse
import copy
import json
import yaml
import textfsm
import io
import collections.abc
import sys
from openpyxl import Workbook
from openpyxl.worksheet.table import Table , TableStyleInfo

class Logging(Exception):
    """
    Exception class handling the exception raised by this script
    """
    def __init__(self, debug=False):
        self.indent_symbol = " "
        self.debug_enabled = debug

    def fatal(self, msg):
        """
        Prints an error message and aborts program execution
        """
        sys.stderr.write(msg+"\n")
        sys.exit(1)

    def warning(self, msg):
        """
        Prints a warning message to stderr
        """
        sys.stderr.write(msg+"\n")
    
    def info(self, msg, level=0):
        """
        Prints a message to stdout
        """
        print('{}'.format(self.indent_symbol*level) + msg)

    def debug(self, msg, level=0):
        """
        Prints a debug message to stdout
        """
        if self.debug_enabled:
            print('{}'.format(self.indent_symbol*level) + msg)

def read_arguments():
    """
    """
    parser = argparse.ArgumentParser("Usage: analysis.py")
    parser.add_argument("--debug", dest="debug", help=" Enable Debug", default=False, required=False)
    parser.add_argument("--base_dir", dest="base_dir", help=" Base Directory", default="./discovery-output/", required=False)
    parser.add_argument("--state_file", dest="state_file", help=" Discovery State File", default="discovery_results.json", required=False)
    parser.add_argument("--config_dir", dest="config_dir", help=" Discovery Config Directory", default="configs", required=False)
    parser.add_argument("--datamodel", dest="datamodel", help=" Parsable Datamodel", default="data_model.yaml", required=False)
    parser.add_argument("--composer", dest="composer", help=" Excel Composer", default="composer.yaml", required=False)
    parser.add_argument("--output", dest="output", help=" Output Excel File Name", default="discovery-analysis-output.xlsx", required=False)
    args = parser.parse_args()
    return args

def load_json(file):
    """
    """
    try:
        with open(file, "r") as file:
            dictionary = json.load(file)
        file.close()
        return(dictionary)
    except:
        LOG.fatal("Error occured while importing %s as YAML input" % file)

def load_yaml(file):
    """
    """
    try:
        with open(file, "r") as file:
            dictionary = yaml.load(file, Loader=yaml.SafeLoader)
        file.close()
        return(dictionary)
    except:
        LOG.fatal("Error occured while importing %s as JSON input" % file)

def analyse_discovered_device_commands(data_model, data):
    # Define dictionary to store analysis result
    command_analysis_result = {}

    # Loop though the list of discovered devides
    for device in data.keys():
        device_network_os = data[device]["ansible_network_os"]
        LOG.info("- Analysing date from \"%s\"" % device)
        LOG.info("- Network OS: %s" % device_network_os, 4)

        # Generate command_parser structure mapping based on data_model and network_os
        command_parser = {}
        for model_catagory in data_model.keys():
            try:
                if device_network_os in data_model[model_catagory]["command"].keys():
                    command = data_model[model_catagory]["command"][device_network_os]
                    command_parser[command] = {}
                    command_parser[command]["fsm_template"] = data_model[model_catagory]["command_parser"][device_network_os]
                    command_parser[command]["variables"] = data_model[model_catagory]["parsable_data"]
                    command_parser[command]["parent"] = model_catagory
            except KeyError:
                # Not a command parser entry in the data model, skipping
                continue
                
        # Analysing command output            
        command_analysis_result[device] = analyse_dicovery_commands(command_parser, data[device]["manual_discovery_commands"])
    
    # Return Analysis result
    return(command_analysis_result)

def analyse_dicovery_commands(parser, data):
    # Define dictionary to store analysed data
    device_config_analysis_result = {}

    # Loop through the list of command results, one command at a time
    for cmd_data in data["results"]:
        command = cmd_data["item"]
        LOG.info("- Analysing output \"%s\"" % command, 4)

        # Check if command execution failed
        if cmd_data["failed"]:
            LOG.info("- Skipping (command execution failed or where ignorred during discovery)", 6)
            continue

        # Check if command is defined in data model
        if command not in parser.keys():
            LOG.info("- Warning, no parser defined for commnand. Skipping.....", 6)
            continue
        
        # Parse command output
        if len(cmd_data["stdout"]) == 1:
            command_output = cmd_data["stdout"][0]
            command_parser = parser[command]["fsm_template"]
        else:
            LOG.info("- Error, an error occured while analysing command output. Skipping.....", 6)
            continue

        parsing_result = parse_output(command_parser, command_output)

        # Record analyis output
        if len(parsing_result) > 0:
            LOG.info("- found %s results from analysis" % len(parsing_result), 6)
            parent_record = parser[command]["parent"]
            device_config_analysis_result[parent_record] = []
            for analysis_entry in parsing_result:
                saved_record = {}

                # Save analysis output
                i = 0
                while i < len(parser[command]["variables"]):
                    saved_record[parser[command]["variables"][i]] = analysis_entry[i]
                    i += 1
                device_config_analysis_result[parent_record].append(saved_record)
        else:
            LOG.info("- No data found during analysis", 6)

    # Record analysis results
    return(device_config_analysis_result)

def analyse_discovered_device_config(data_model, data):
    # Define dictionary to store analysis result
    config_analysis_result = {}

    # Loop though the list of discovered devides
    for device in data.keys():
        device_network_os = data[device]["ansible_network_os"]
        device_config_file = config_directory + device + ".txt"
        LOG.info("- Analysing date from \"%s\"" % device)
        LOG.info("- Network OS: %s" % device_network_os, 4)

        # Generate command_parser structure mapping based on data_model and network_os
        config_parser = {}
        for model_catagory in data_model.keys():
            try:
                if device_network_os in data_model[model_catagory]["config_parser"].keys():
                    config_parser[model_catagory] = {}
                    config_parser[model_catagory]["fsm_template"] = data_model[model_catagory]["config_parser"][device_network_os]
                    config_parser[model_catagory]["variables"] = data_model[model_catagory]["parsable_data"]
            except KeyError:
                # Not a config parser entry in the data model, skipping
                continue

        # Skip if no config parser found for the respective Network OS
        if len(config_parser) == 0:
            LOG.info("- Warning, no config parser defined for %s. Skipping....." % device_network_os, 6)
            continue

        # # Analysing config          
        config_analysis_result[device] = analyse_dicovery_config(config_parser, device_config_file)

    # Return Analysis result
    return(config_analysis_result)

def analyse_dicovery_config(parser, config_file):
    # Define dictionary to store analysed data
    device_config_analysis_result = {}

    # Read config file
    config = ""
    try:
        with open(config_file, "r") as file:
            config = file.read()
        file.close()
    except:
        LOG.fatal("Error occured while reading config file \'%s\'" % config_file)
        
    # Loop through the list of config elements that needs to be parsable_data
    for cfg_element in parser:
        LOG.info("- Analysing Config Element \"%s\"" % cfg_element, 4)

        # Parse config
        parsing_result = parse_output(parser[cfg_element]["fsm_template"], config)

        # Record analyis output
        if len(parsing_result) > 0:
            LOG.info("- found %s results from analysis" % len(parsing_result), 6)
            device_config_analysis_result[cfg_element] = []
            for analysis_entry in parsing_result:
                saved_record = {}

                # Save analysis output
                i = 0
                while i < len(parser[cfg_element]["variables"]):
                    saved_record[parser[cfg_element]["variables"][i]] = analysis_entry[i]
                    i += 1
                device_config_analysis_result[cfg_element].append(saved_record)
        else:
            LOG.info("- No data found during analysis", 6)

    # Record analysis results
    return(device_config_analysis_result)

def parse_output(parser_file, output):
    # Read parser syntax
    try:
        parser_syntax = ""
        file = open("parsers/" + parser_file, "r")
        for line in file:
            parser_syntax = parser_syntax + line
        file.close()
    except:
        LOG.fatal("Error occured while importing %s as TextFSM parser schema" % parser_file)

    # Parse command output
    fsm_Result = textfsm.TextFSM(io.StringIO(parser_syntax))
    parser_result = fsm_Result.ParseText(output)

    # Return data
    return(parser_result)

def update_dict(d, u):
    for k, v in u.items():
        if isinstance(v, collections.abc.Mapping):
            d[k] = update_dict(d.get(k, {}), v)
        else:
            d[k] = v
    return d

def extract_sheet_data(composer, sheet, data):
    extracted_data = []

    # Extract data source
    try:
        data_source = composer[sheet]["__data_source__"]
    except KeyError:
        LOG.info("- ERROR, data source missing from composer. Skipping.....", 4)


    # loop through discovered data on a per-device basis
    for device in data:
        device_extracted_data = []

        # Skip devices, where no data are available
        if data_source not in data[device].keys():
            continue
        else:
            LOG.info("- Extracting data from device \"%s\"" % device, 4)

        # Loop through the discovered data
        for entry in data[device][data_source]:
            data_entries = {}

            # Loop through the sheet keys and gather data
            for header_name in composer[sheet].keys():
                data_field_name = composer[sheet][header_name]

                # Skip data source key,  as this one is not a sheet header
                if header_name == "__data_source__":
                    continue

                # Handle special data field named __device_name__
                if data_field_name == "__device_name__":
                    data_entries["device_name"] = device
                    continue

                # Add data
                try:
                    data_entries[header_name] = entry[data_field_name]
                except KeyError:
                    # Data field found during discovery
                    data_entries[header_name] = ""

            device_extracted_data.append(data_entries)

        # Aggregate extracted data
        LOG.info("- found %s datasets" % len(device_extracted_data), 6)
        extracted_data = extracted_data + device_extracted_data

    return(extracted_data)

def create_worksheet(excel_workbook,sheet_name,headers,extracted_data):
    """
    """
    sheet = excel_workbook.create_sheet(title = sheet_name)
    if len(headers) > 0:
        for i in range(0,len(headers)):
            sheet.cell(column=i+1, row=1 , value = headers[i])
        row_id = 2

        for element in extracted_data:
            for i in range(0,len(headers)):
                try:
                    if len(str(element[headers[i]])) == 1:
                        sheet.cell(column = i+1, row = row_id, value = "{0}".format(str(element[headers[i]][0])))
                    else:
                        sheet.cell(column = i+1, row = row_id, value = "{0}".format(",".join(str(element[headers[i]]))))
                    sheet.cell(column = i+1, row = row_id, value = "{0}".format(str(element[headers[i]])))
                except:
                    sheet.cell(column = i+1, row = row_id, value = "")
            row_id = row_id + 1
        if len(headers) > 26:  ### Super Ugly !! Fix That
            table_cells = "A1:A" + chr(64+len(headers)-26)
        else:
            table_cells = "A1:" + chr(64+len(headers)) + str(row_id-1)
        style = TableStyleInfo(name = "TableStyleMedium9" , showRowStripes="True" )
        table = Table(displayName = sheet_name , ref = table_cells)
        table.tableStyleInfo = style
        sheet.add_table(table)
    return excel_workbook

def create_workbook(composer, data):
    excel_workbook = Workbook()
    position = 0
    for sheet in composer.keys():
        LOG.info("- Composing sheet \"%s\"" % sheet, 2)
        # Define sheet composer
        sheet_header = []

        # Extract data source
        try:
            data_source = composer[sheet]["__data_source__"]
        except KeyError:
            LOG.info("- Error, data source missing from composer. Skipping.....", 6)
            continue

        # Construct sheet headers and associated data fields
        LOG.info("- Constructing sheet headers", 4)
        for key in composer[sheet].keys():
            # Skip data source key,  as this one is not a sheet header
            if key == "__data_source__":
                continue

            sheet_header.append(key)

        # Extract sheet data
        sheet_data = extract_sheet_data(composer, sheet, data)

        # Create worksheet
        LOG.info("- Creating sheet", 4)
        excel_workbook = create_worksheet(excel_workbook, sheet, sheet_header, sheet_data)

    return excel_workbook

def main():
    LOG.info("+ ---------------------------------------------------")
    LOG.info("+ Stage 0: Starting Analysis")
    LOG.info("+ ---------------------------------------------------")
    # Read parsable data model
    LOG.info("- Importing parsable data model (%s)" % data_model_file)
    data_model = load_yaml(data_model_file)

    # Read discovery output
    LOG.info("- Importing Discovered Statefile (%s)" % state_file)
    discovery_output = load_json(state_file)

    LOG.info("")
    LOG.info("+ ---------------------------------------------------")
    LOG.info("+ Stage 1: Analysing Discovery Data (command output)")
    LOG.info("+ ---------------------------------------------------")
    # Analyse discovery configs (command output)
    command_analysis_result = analyse_discovered_device_commands(data_model, discovery_output)

    LOG.info("")
    LOG.info("+ ---------------------------------------------------")
    LOG.info("+ Stage 2: Analysing Discovery Data (configs)")
    LOG.info("+ ---------------------------------------------------")
    # Analyse discovery data (configs)
    config_analysis_result = analyse_discovered_device_config(data_model, discovery_output)

    # Combine analysis results
    LOG.debug("- Combining analysis results")
    analysis_result = copy.deepcopy(command_analysis_result)
    analysis_result = update_dict(analysis_result, config_analysis_result)

    LOG.info("")
    LOG.info("+ ---------------------------------------------------")
    LOG.info("+ Stage 3: Writing Analysis Results to disk")
    LOG.info("+ ---------------------------------------------------")

    # Importing Excel Composer
    LOG.info("- Importing Output Composer (%s)" % composer_file)
    composer = load_yaml(composer_file)

    # Compose Workbook
    LOG.info("- Composing Workbook")
    excel_workbook = create_workbook(composer, analysis_result)

    # Write Workbook to disc
    LOG.info("- Writing Workbook to disk")
    excel_workbook.save(output_file)

if __name__ == '__main__':
    # Read arguments
    args = read_arguments()

    # Define base variables
    base_dir = args.base_dir
    data_model_file = args.datamodel
    composer_file = args.composer
    output_file = args.output
    state_file = base_dir + args.state_file
    config_directory = base_dir + args.config_dir + "/"

    # Define logging
    LOG = Logging(debug=args.debug)

    # Initialize analysis
    main()
